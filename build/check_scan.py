"""The scanner + writer through the shipped path, over synthetic projects.

No ArcGIS Pro exists on the dev PC, so the fixtures are built here in the
shape Pro writes: zips of CIM JSON. Every assertion reads the FILE that comes
out (the workbook, the diagnostics, the bundle), not an in-memory shortcut.

    python build\\check_scan.py
"""
import io
import json
import sys
import zipfile

from _checklib import Checker, patch, scripts_path, temp_dir

scripts_path()

import aprx_scan  # noqa: E402
import scan_output  # noqa: E402
from aprx_scan import (classify_environment, describe_source, parse_connection_string,  # noqa: E402
                       parse_service_url, redact_connection_string, redact_document, run_scan)
from events import Events  # noqa: E402
from scan_output import save, summary_lines, table_rows, write_bundle  # noqa: E402

PROD = "https://gis-prod.example.org/server/rest/services/TSMIS/lrs_tsmis/FeatureServer"
TEST = "https://gis-test.example.org/server/rest/services/TSMIS_QA/lrs_tsmis/FeatureServer"
SECRET = "hunter2-base64=="


def conn(text, factory="FeatureService", dataset="3"):
    return {"type": "CIMStandardDataConnection", "workspaceConnectionString": text,
            "workspaceFactory": factory, "dataset": dataset, "datasetType": "esriDTFeatureClass"}


def layer(name, connection, ctype="CIMFeatureLayer"):
    return {"type": ctype, "name": name, "uRI": f"CIMPATH=map/{name.lower()}.json",
            "featureTable": {"type": "CIMFeatureTable", "dataConnection": connection}}


def map_doc(name, layers, extra=None):
    doc = {"type": "CIMMapDocument", "version": "3.2.0", "build": 49743,
           "mapDefinition": {"type": "CIMMap", "name": name,
                             "layers": [ly["uRI"] for ly in layers if "uRI" in ly]},
           "layerDefinitions": layers}
    doc.update(extra or {})
    return doc


def write_aprx(path, members):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, payload in members.items():
            zf.writestr(name, payload if isinstance(payload, bytes) else json.dumps(payload))
    path.write_bytes(buf.getvalue())


def build_fixtures(root):
    # Alpha: a prod feature-service layer on a branch, a file-gdb layer with no
    # version, a standalone table on an SDE connection (dev server, password),
    # a group layer, and a second map. GISProject.json + a binary thumbnail too.
    alpha_map = map_doc("Highways", [
        layer("Highways", conn(f"URL={PROD};VERSION=OWNER.Branch_A;VERSIONGUID={{ABC-123}}")),
        layer("Counties", conn("DATABASE=C:\\data\\base.gdb", factory="FileGDB", dataset="Counties")),
        {"type": "CIMGroupLayer", "name": "Reference", "layers": ["CIMPATH=map/counties.json"]},
    ])
    alpha_table = {"type": "CIMStandaloneTable", "name": "Inventory",
                   "dataConnection": conn(
                       "ENCRYPTED_PASSWORD=" + SECRET + ";SERVER=dbhost-dev;INSTANCE=sde:sqlserver:dbhost-dev;"
                       "DBCLIENT=sqlserver;DATABASE=tsmis;USER=reader;VERSION=sde.DEFAULT;"
                       "AUTHENTICATION_MODE=DBMS", factory="SDE", dataset="tsmis.dbo.Inventory")}
    alpha_map2 = map_doc("Ramps", [layer("Ramps", conn(f"URL={PROD};VERSION=owner.branch_a"))],
                         extra={"tableDefinitions": [alpha_table],
                                "serviceAccount": {"password": "plain-secret", "user": "svc"}})
    write_aprx(root / "Alpha.aprx", {
        "GISProject.json": {"type": "CIMGISProject", "version": "3.2.0", "projectItems": []},
        "Maps/m1.json": alpha_map, "Maps/m2.json": alpha_map2,
        "Index/thumb.png": b"\x89PNG\r\n\x1a\n\x00binary",
    })
    # Beta: two layers on two versions, on the test host.
    write_aprx(root / "Beta.aprx", {"Maps/m.json": map_doc("Map", [
        layer("A", conn(f"URL={TEST};VERSION=SDE.DEFAULT")),
        layer("B", conn(f"URL={TEST};VERSION=KELLY.QA_2026")),
    ])})
    # Empty: JSON but no connections. Corrupt: not a zip.
    write_aprx(root / "Empty.aprx", {"GISProject.json": {"type": "CIMGISProject"}})
    (root / "Corrupt.aprx").write_bytes(b"this is not a zip archive")
    # Nested project, a .backups copy that must be skipped, a stray file, and
    # a plain-JSON layer file (only read when asked).
    (root / "Nested" / "sub").mkdir(parents=True)
    (root / "Nested" / ".backups").mkdir()
    write_aprx(root / "Nested" / "sub" / "Gamma.aprx",
               {"Maps/m.json": map_doc("Map", [layer("G", conn(f"URL={PROD};VERSION=OWNER.Gamma"))])})
    write_aprx(root / "Nested" / ".backups" / "Gamma.aprx",
               {"Maps/m.json": map_doc("Map", [layer("G", conn(f"URL={PROD};VERSION=OWNER.Old"))])})
    (root / "notes.txt").write_text("not a project", encoding="utf-8")
    (root / "Layer.lyrx").write_text(json.dumps({
        "type": "CIMLayerDocument", "version": "3.2.0",
        "layerDefinitions": [layer("Solo", conn(f"URL={PROD};VERSION=OWNER.Lyrx"))]}),
        encoding="utf-8")


def test_helpers(c):
    print("connection-string helpers:")
    props = parse_connection_string("URL=https://h/x?a=b=c;VERSION=sde.DEFAULT; user = me ")
    c.check("keys upper-cased, values keep '='", props == {"URL": "https://h/x?a=b=c",
                                                          "VERSION": "sde.DEFAULT", "USER": "me"}, props)
    red = redact_connection_string("ENCRYPTED_PASSWORD=abc==;SERVER=s;TOKEN=t;VERSION=v")
    c.check("passwords/tokens removed, the rest kept",
            red == "ENCRYPTED_PASSWORD=<removed>;SERVER=s;TOKEN=<removed>;VERSION=v", red)
    c.check("service URL wins as the source", describe_source({"URL": "https://h/FeatureServer?f=json",
                                                               "SERVER": "x"}) == "https://h/FeatureServer")
    c.check("server · instance · database otherwise",
            describe_source({"SERVER": "s", "INSTANCE": "i", "DATABASE": "d"}) == "s · i · d")

    print("service URL parsing + environment:")
    u = parse_service_url(PROD)
    c.check("the real TSMIS prod URL -> host / site / folder / service / Prod",
            u == {"host": "gis-prod.example.org", "site": "server", "folder": "TSMIS",
                  "service": "lrs_tsmis", "environment": "Prod"}, u)
    u = parse_service_url("https://gis-dev.example.org/server/rest/services/A/B/svc/MapServer/3")
    c.check("nested folders join with '/', server type + layer id dropped, Dev",
            (u["folder"], u["service"], u["environment"]) == ("A/B", "svc", "Dev"), u)
    u = parse_service_url("https://services.arcgis.com/abc123/arcgis/rest/services/Roads/FeatureServer")
    c.check("a hosted layer: no folder, no environment (nothing to read it from)",
            (u["site"], u["folder"], u["service"], u["environment"]) == ("abc123/arcgis", "", "Roads", ""), u)
    c.check("garbage -> blanks", parse_service_url("not a url")["service"] == "" and parse_service_url("")["host"] == "")
    c.check("environment from a folder token when the host has none",
            parse_service_url("https://gis.example.org/server/rest/services/TSMIS_DEV/svc/FeatureServer")["environment"] == "Dev")
    c.check("classify: test host / uat / plain", (classify_environment("gis-test.example.org"),
                                                   classify_environment("x", "uat-site"), classify_environment("tsmis"))
            == ("Test", "Test", ""))

    print("document redaction:")
    doc = redact_document({"a": {"password": "p", "Token": "t", "note": "PASSWORD=x;SERVER=s"},
                           "list": [{"connectionString": "USER=u;ENCRYPTED_PASSWORD=zz"}], "keep": "URL=x;VERSION=v"})
    c.check("secret keys and secrets inside strings removed, the rest intact",
            doc == {"a": {"password": "<removed>", "Token": "<removed>", "note": "PASSWORD=<removed>;SERVER=s"},
                    "list": [{"connectionString": "USER=u;ENCRYPTED_PASSWORD=<removed>"}], "keep": "URL=x;VERSION=v"}, doc)


def test_scan(c, tmp):
    root = tmp / "projects"
    root.mkdir()
    build_fixtures(root)
    print("recursive scan of the fixtures:")
    lines = []
    res = run_scan(root, recursive=True, events=Events(on_log=lines.append))
    by = {p.path.name: p for p in res.projects}
    c.check("finds the 4 top-level + 1 nested .aprx (not the .backups copy, not .lyrx)",
            sorted(by) == ["Alpha.aprx", "Beta.aprx", "Corrupt.aprx", "Empty.aprx", "Gamma.aprx"], sorted(by))
    c.check("one .backups folder skipped", res.skipped_dirs == 1, res.skipped_dirs)
    a = by["Alpha.aprx"]
    c.check("Alpha ok", a.status == "ok", (a.status, a.message))
    c.check("Alpha versions de-duplicate case-insensitively, first spelling kept",
            a.versions() == ["OWNER.Branch_A", "sde.DEFAULT"], a.versions())
    c.check("Alpha environments: Prod (service) + Dev (SDE server), distinct",
            a.environments() == ["Prod", "Dev"], a.environments())
    c.check("Alpha service folders", a.service_folders() == ["TSMIS"], a.service_folders())
    c.check("Alpha maps", a.maps == ["Highways", "Ramps"], a.maps)
    c.check("Alpha has 4 connections (3 layers + 1 table)", len(a.connections) == 4, len(a.connections))
    hw = next(x for x in a.connections if x.layer == "Highways")
    c.check("layer context: name/type/map/factory/guid", (hw.layer_type, hw.map, hw.factory, hw.version_guid)
            == ("FeatureLayer", "Highways", "FeatureService", "{ABC-123}"), hw)
    c.check("service fields on the connection", (hw.host, hw.folder, hw.service, hw.environment)
            == ("gis-prod.example.org", "TSMIS", "lrs_tsmis", "Prod"), hw)
    c.check("connection type label", hw.connection_type == "Feature service")
    inv = next(x for x in a.connections if x.layer == "Inventory")
    c.check("standalone table context + SDE source + env from SERVER", (inv.layer_type, inv.source, inv.version, inv.environment)
            == ("StandaloneTable", "dbhost-dev · sde:sqlserver:dbhost-dev · tsmis", "sde.DEFAULT", "Dev"), inv)
    c.check("password redacted in the connection", SECRET not in inv.connection and "<removed>" in inv.connection)
    c.check("json path recorded", hw.json_path == "layerDefinitions[0].featureTable.dataConnection", hw.json_path)
    c.check("members + types recorded for diagnostics",
            len(a.members) == 4 and sum(1 for m in a.members if m["json"]) == 3
            and a.types_seen.get("CIMFeatureLayer") == 3 and a.members[3].get("head", "").startswith("89504e47"),
            (a.members, a.types_seen))
    c.check("no documents kept in a normal scan", a.documents == [])
    b = by["Beta.aprx"]
    c.check("Beta two versions, Test", b.versions() == ["SDE.DEFAULT", "KELLY.QA_2026"] and b.environments() == ["Test"],
            (b.versions(), b.environments()))
    c.check("Empty -> no_connections with a count", by["Empty.aprx"].status == "no_connections"
            and "1 files inside, 1 JSON" in by["Empty.aprx"].message, by["Empty.aprx"].message)
    c.check("Corrupt -> error (not a zip)", by["Corrupt.aprx"].status == "error"
            and "not a zip" in by["Corrupt.aprx"].message, by["Corrupt.aprx"].message)
    c.check("nested Gamma found", by["Gamma.aprx"].versions() == ["OWNER.Gamma"])
    c.check("per-file log lines carry the environment",
            any(l.startswith("  Alpha.aprx: OWNER.Branch_A, sde.DEFAULT  [Prod, Dev]") for l in lines)
            and any("Corrupt.aprx: Error" in l for l in lines), lines)
    tally = res.version_tally()
    c.check("version tally keyed by (environment, version); spelling folds to the first sighting",
            tally[("Prod", "OWNER.Branch_A")]["layers"] == 2 and tally[("Dev", "sde.DEFAULT")]["projects"] == ["Alpha.aprx"]
            and tally[("Test", "sde.DEFAULT")]["projects"] == ["Beta.aprx"]
            and ("Test", "SDE.DEFAULT") not in tally, tally)
    c.check("counts", res.counts() == {"ok": 3, "no_versions": 0, "no_connections": 1, "error": 1,
                                       "total": 5, "cloud_only": 0}, res.counts())

    print("save through the writer:")
    workbook, diagnostics = save(res, tmp / "out")
    from openpyxl import load_workbook
    wb = load_workbook(str(workbook))
    c.check("four sheets", wb.sheetnames == ["Projects", "Layers", "Versions", "Scan"], wb.sheetnames)
    rows = list(wb["Projects"].iter_rows(values_only=True))
    c.check("Projects header", rows[0][:5] == ("Project", "Location", "Status", "Environments", "Versions"), rows[0])
    c.check("Projects sorted by folder then name; Alpha first",
            [r[0] for r in rows[1:]] == ["Alpha.aprx", "Beta.aprx", "Corrupt.aprx", "Empty.aprx", "Gamma.aprx"],
            [r[0] for r in rows[1:]])
    c.check("Alpha row: status, environments, versions, service folders",
            rows[1][2:6] == ("OK", "Prod | Dev", "OWNER.Branch_A | sde.DEFAULT", "TSMIS"), rows[1])
    lay = list(wb["Layers"].iter_rows(values_only=True))
    c.check("Layers header carries Environment/Host/Service folder/Service",
            lay[0][5:9] == ("Environment", "Host", "Service folder", "Service"), lay[0])
    c.check("Layers rows = every connection (4+2+1)", len(lay) == 8, len(lay))
    hw_row = next(r for r in lay[1:] if r[2] == "Highways")
    c.check("Highways row: Prod / host / TSMIS / lrs_tsmis / version",
            hw_row[5:10] == ("Prod", "gis-prod.example.org", "TSMIS", "lrs_tsmis", "OWNER.Branch_A"), hw_row)
    blob = json.dumps([list(r) for r in lay], default=str)
    c.check("no password anywhere in the Layers sheet", SECRET not in blob)
    ver = list(wb["Versions"].iter_rows(values_only=True))
    c.check("Versions sheet keyed by environment + version",
            ver[0] == ("Environment", "Version", "Projects", "Layers", "Project files")
            and any(r[:4] == ("Prod", "OWNER.Branch_A", 1, 2) for r in ver[1:])
            and any(r[:3] == ("Test", "sde.DEFAULT", 1) for r in ver[1:]), ver)
    diag = json.loads(diagnostics.read_text(encoding="utf-8"))
    c.check("diagnostics carry members/types/connections/environment and no password",
            len(diag["files"]) == 5 and diag["files"][0]["types_seen"]
            and diag["files"][0]["connections"][0]["environment"] == "Prod"
            and SECRET not in diagnostics.read_text(encoding="utf-8"))
    c.check("summary lines name the environment and the workbook",
            summary_lines(res, workbook)[0].startswith("Read 5 file(s)")
            and "OWNER.Branch_A (Prod): 1 project" in summary_lines(res, workbook)[1]
            and str(workbook) in summary_lines(res, workbook)[-1], summary_lines(res, workbook))
    tr = table_rows(res)
    c.check("table rows are JSON-safe, ordered, with environments",
            tr[0]["name"] == "Alpha.aprx" and tr[0]["environments"] == ["Prod", "Dev"] and json.dumps(tr))

    print("options:")
    flat = run_scan(root, recursive=False)
    c.check("non-recursive: top level only", sorted(p.path.name for p in flat.projects)
            == ["Alpha.aprx", "Beta.aprx", "Corrupt.aprx", "Empty.aprx"])
    extra = run_scan(root, recursive=False, include_map_layer_files=True)
    ly = next((p for p in extra.projects if p.path.suffix == ".lyrx"), None)
    c.check(".lyrx read as a layer file when asked", ly is not None and ly.kind == "layer file"
            and ly.versions() == ["OWNER.Lyrx"], ly)
    calls = {"n": 0}

    def cancel_after_first():
        calls["n"] += 1
        return calls["n"] > 1
    cancelled = run_scan(root, recursive=True, events=Events(is_cancelled=cancel_after_first))
    c.check("cancel stops after the current file", cancelled.cancelled and len(cancelled.projects) == 1)
    try:
        run_scan(tmp / "missing")
        c.check("missing folder raises ScanError", False)
    except aprx_scan.ScanError:
        c.check("missing folder raises ScanError", True)

    print("diagnostics bundle:")
    logs = tmp / "logs"
    logs.mkdir()
    (logs / "tsmis-gui.log").write_text("log line\n", encoding="utf-8")
    deep = run_scan(root, recursive=True, keep_documents=True)
    da = next(p for p in deep.projects if p.path.name == "Alpha.aprx")
    c.check("bundle mode keeps every JSON document of Alpha (3)", [m for m, _t in da.documents]
            == ["GISProject.json", "Maps/m1.json", "Maps/m2.json"], da.documents and [m for m, _t in da.documents])
    m2 = json.loads(next(t for m, t in da.documents if m == "Maps/m2.json"))
    c.check("kept documents are redacted (connection password + a plain password key)",
            SECRET not in json.dumps(m2) and m2["serviceAccount"]["password"] == "<removed>"
            and m2["serviceAccount"]["user"] == "svc" and "<removed>" in
            m2["tableDefinitions"][0]["dataConnection"]["workspaceConnectionString"])
    with patch(scan_output, "LOG_DIR", logs):
        bundle = write_bundle(deep, tmp / "bundle" / "diag.zip", workbook)
    with zipfile.ZipFile(bundle) as zf:
        names = zf.namelist()
        c.check("bundle holds summary + workbook + raw docs per project + the log",
                "summary.json" in names and "branch_versions.xlsx" in names and "logs/tsmis-gui.log" in names
                and "raw/001 Alpha/Maps__m2.json.json" in names and "raw/002 Beta/Maps__m.json.json" in names
                and not any(n.startswith("raw/003") and n.endswith(".json") and "Corrupt" in n for n in names), names)
        payload = json.loads(zf.read("summary.json"))
        c.check("summary.json lists the kept documents and carries no password",
                payload["files"][0]["documents_kept"] == ["GISProject.json", "Maps/m1.json", "Maps/m2.json"]
                and SECRET not in zf.read("summary.json").decode("utf-8"))
        raw = zf.read("raw/001 Alpha/Maps__m1.json.json").decode("utf-8")
        c.check("a raw document round-trips as JSON with the real structure",
                json.loads(raw)["layerDefinitions"][0]["name"] == "Highways")
        c.check("no password anywhere in the bundle",
                all(SECRET not in zf.read(n).decode("utf-8", "replace") for n in names if n.endswith(".json")))


def main():
    c = Checker()
    test_helpers(c)
    with temp_dir("tsmis_scan_") as tmp:
        test_scan(c, tmp)
    raise SystemExit(c.summary())


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    main()
