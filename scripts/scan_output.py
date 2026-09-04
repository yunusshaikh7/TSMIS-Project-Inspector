"""Turn a ScanResult into what people read: the Excel workbook, the
diagnostics JSON beside it, the summary lines, the rows the GUI table shows,
and — on request — the diagnostics BUNDLE for the maintainer. Console-free.

The workbook is the deliverable (Projects / Layers / Versions / Scan sheets).
The diagnostics file records what the reader SAW inside every file (archive
members, CIM types, every connection string with its JSON path). The bundle
adds every document itself, secrets removed, plus the workbook and the app
log — one zip that is enough to teach the reader a real project library's
layout without the projects ever leaving the PC.
"""
import json
import logging
import re
import zipfile
from pathlib import Path

from paths import LOG_DIR, new_run_dir
from version import APP_NAME, __version__

log = logging.getLogger("tsmis.scan")

WORKBOOK_NAME = "branch_versions.xlsx"
DIAGNOSTICS_NAME = "diagnostics.json"

PROJECT_COLUMNS = ("Project", "Location", "Status", "Environments", "Versions", "Service folders",
                   "Services", "Layers with data", "Maps", "Type", "Cloud-only", "Size (KB)",
                   "Modified", "Note")
LAYER_COLUMNS = ("Project", "Map", "Layer", "Layer type", "Connection type", "Environment", "Host",
                 "Service folder", "Service", "Version", "Version GUID", "Service / workspace",
                 "Dataset", "Connection string (passwords removed)", "Found in")
VERSION_COLUMNS = ("Environment", "Version", "Projects", "Layers", "Project files")
_MAX_COL_WIDTH = 70
_SAFE_NAME = re.compile(r"[^A-Za-z0-9._ -]+")


def save(result, out_dir=None):
    """Write the workbook + diagnostics into `out_dir` (a fresh run folder by
    default). Returns (workbook_path, diagnostics_path)."""
    out_dir = Path(out_dir) if out_dir else new_run_dir(result.started)
    out_dir.mkdir(parents=True, exist_ok=True)
    workbook = out_dir / WORKBOOK_NAME
    diagnostics = out_dir / DIAGNOSTICS_NAME
    write_diagnostics(result, diagnostics)
    write_workbook(result, workbook)
    log.info("scan results saved: %s", workbook)
    return workbook, diagnostics


def _sorted_projects(result):
    return sorted(result.projects, key=lambda p: (str(p.path.parent).lower(), p.path.name.lower()))


def project_row(p):
    return (p.path.name, str(p.path.parent), p.status_text, " | ".join(p.environments()),
            " | ".join(p.versions()), " | ".join(p.service_folders()), " | ".join(p.sources()),
            sum(1 for c in p.connections if c.source or c.version), ", ".join(p.maps), p.kind,
            "yes" if p.cloud_only else "", round(p.size / 1024, 1), p.modified, p.message)


def layer_rows(p):
    for c in p.connections:
        yield (p.path.name, c.map, c.layer, c.layer_type, c.connection_type, c.environment, c.host,
               c.folder, c.service, c.version, c.version_guid, c.source, c.dataset, c.connection,
               f"{c.member} · {c.json_path}")


def write_workbook(result, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_font = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="DDE7F3")

    def sheet(title, columns, rows, first=False):
        ws = wb.active if first else wb.create_sheet()
        ws.title = title
        ws.append(list(columns))
        widths = [len(c) for c in columns]
        for row in rows:
            ws.append(list(row))
            for i, v in enumerate(row):
                widths[i] = max(widths[i], min(_MAX_COL_WIDTH, len(str(v)) if v is not None else 0))
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w + 2
        for cell in ws[1]:
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        return ws

    projects = _sorted_projects(result)
    sheet("Projects", PROJECT_COLUMNS, (project_row(p) for p in projects), first=True)
    sheet("Layers", LAYER_COLUMNS, (r for p in projects for r in layer_rows(p)))
    tally = result.version_tally()
    sheet("Versions", VERSION_COLUMNS,
          ((env, ver, len(t["projects"]), t["layers"], ", ".join(t["projects"]))
           for (env, ver), t in sorted(tally.items(), key=lambda kv: (-len(kv[1]["projects"]), kv[0]))))
    sheet("Scan", ("Item", "Value"), scan_facts(result))
    wb.save(str(path))


def scan_facts(result):
    c = result.counts()
    return [
        ("App", f"{APP_NAME} v{__version__}"),
        ("Folder scanned", str(result.root)),
        ("Subfolders included", "yes" if result.recursive else "no"),
        ("Map/layer files (.mapx/.lyrx) included", "yes" if result.include_map_layer_files else "no"),
        ("Started", result.started.strftime("%Y-%m-%d %H:%M:%S")),
        ("Finished", result.finished.strftime("%Y-%m-%d %H:%M:%S") if result.finished else ""),
        ("Files read", c["total"]),
        ("With a version", c["ok"]),
        ("Without a version", c["no_versions"]),
        ("Without data connections", c["no_connections"]),
        ("Errors", c["error"]),
        ("OneDrive cloud-only files (downloaded to read)", c["cloud_only"]),
        (".backups folders skipped", result.skipped_dirs),
        ("Folders that could not be listed", result.unreadable_dirs),
        ("Cancelled before the end", "yes" if result.cancelled else "no"),
    ]


def diagnostics_payload(result):
    return {
        "app": f"{APP_NAME} v{__version__}",
        "scan": {k: str(v) for k, v in scan_facts(result)},
        "files": [{
            "path": str(p.path), "kind": p.kind, "status": p.status, "message": p.message,
            "size": p.size, "cloud_only": p.cloud_only, "seconds": round(p.seconds, 3),
            "maps": p.maps, "members": p.members, "types_seen": p.types_seen,
            "documents_kept": [m for m, _t in p.documents],
            "connections": [{
                "map": c.map, "layer": c.layer, "layer_type": c.layer_type,
                "factory": c.factory, "version": c.version, "version_guid": c.version_guid,
                "environment": c.environment, "host": c.host, "folder": c.folder,
                "service": c.service, "source": c.source, "dataset": c.dataset,
                "connection": c.connection, "member": c.member, "json_path": c.json_path,
            } for c in p.connections],
        } for p in result.projects],
    }


def write_diagnostics(result, path):
    Path(path).write_text(json.dumps(diagnostics_payload(result), indent=1), encoding="utf-8")


def write_bundle(result, zip_path, workbook=None):
    """One zip for the maintainer: summary.json (the diagnostics), the
    workbook, every kept document as raw/<n> <project>/<member>.json (secrets
    removed at read time), and the app's log files. Returns zip_path."""
    zip_path = Path(zip_path)
    zip_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("summary.json", json.dumps(diagnostics_payload(result), indent=1))
        if workbook and Path(workbook).is_file():
            zf.write(workbook, WORKBOOK_NAME)
        for n, p in enumerate(_sorted_projects(result), 1):
            folder = f"raw/{n:03d} {_SAFE_NAME.sub('_', p.path.stem)[:60]}"
            for member, text in p.documents:
                name = _SAFE_NAME.sub("_", member.replace("\\", "/").replace("/", "__"))
                zf.writestr(f"{folder}/{name}.json", text)
        for log_file in sorted(Path(LOG_DIR).glob("tsmis*.log*")) if Path(LOG_DIR).is_dir() else []:
            try:
                zf.write(log_file, f"logs/{log_file.name}")
            except OSError as e:
                log.warning("bundle: could not add %s (%s)", log_file, e)
    log.info("diagnostics bundle written: %s", zip_path)
    return zip_path


def summary_lines(result, workbook=None):
    c = result.counts()
    lines = [f"Read {c['total']} file(s) under {result.root}: {c['ok']} with a version, "
             f"{c['no_versions']} without a version, {c['no_connections']} without data "
             f"connections, {c['error']} error(s)." + (" (cancelled)" if result.cancelled else "")]
    tally = result.version_tally()
    if tally:
        parts = [f"{ver}{' (' + env + ')' if env else ''}: {len(t['projects'])} project"
                 f"{'s' if len(t['projects']) != 1 else ''}"
                 for (env, ver), t in sorted(tally.items(), key=lambda kv: (-len(kv[1]["projects"]), kv[0]))]
        lines.append("Versions found: " + "; ".join(parts))
    if workbook:
        lines.append(f"Workbook saved: {workbook}")
    return lines


def table_rows(result):
    """JSON-safe rows for the GUI results table."""
    return [{
        "name": p.path.name, "folder": str(p.path.parent), "status": p.status,
        "status_text": p.status_text, "environments": p.environments(), "versions": p.versions(),
        "layers": len(p.connections), "message": p.message,
    } for p in _sorted_projects(result)]
