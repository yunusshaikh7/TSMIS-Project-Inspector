"""The comprehensive runtime self-test — ONE body, two callers, so the exact
shipped exe (`gui_main --self-test`, the build.ps1 -SelfTest release gate)
and the dev venv (build/check_self_test.py) prove the same thing:

  1. every app module imports;
  2. the bundled ui/ assets are present;
  3. a synthetic .aprx (a zip of CIM JSON, built here) scans through the
     REAL run_scan -> save path and the workbook + diagnostics come out
     with the expected version in them (openpyxl round-trip included);
  4. the GUI bridge builds and answers get_initial_state;
  5. best-effort: a hidden WebView2 window cycles through the real JS bridge
     (the only sub-check allowed to skip — a headless box can't show one).

Returns 0 on success and RAISES on any mandatory failure. `emit` is the line
sink (default print). A diagnostic driver, never imported by the engine.
"""
import io
import json
import shutil
import sys
import tempfile
import threading
import time
import zipfile
from pathlib import Path

APP_MODULES = ("version", "paths", "settings", "logging_setup", "events", "aprx_scan",
               "scan_output", "updater", "cli", "gui_win32", "gui_worker", "gui_api",
               "gui_main", "self_test")

FEATURE_SERVICE = "https://gis.example.org/server/rest/services/TSMIS/FeatureServer"


def synthetic_aprx(path, version="OWNER.Branch_A", layer="Highways", map_name="Map"):
    """A minimal .aprx: a zip holding CIM JSON in the shape Pro writes — a map
    document whose layer definitions carry feature-service connections."""
    layer_def = {
        "type": "CIMFeatureLayer", "name": layer, "uRI": "CIMPATH=map/highways.json",
        "featureTable": {
            "type": "CIMFeatureTable", "displayField": "NAME",
            "dataConnection": {
                "type": "CIMStandardDataConnection",
                "workspaceConnectionString": f"URL={FEATURE_SERVICE};VERSION={version}",
                "workspaceFactory": "FeatureService", "dataset": "3",
                "datasetType": "esriDTFeatureClass"}}}
    map_doc = {"type": "CIMMapDocument", "version": "3.2.0", "build": 49743,
               "mapDefinition": {"type": "CIMMap", "name": map_name,
                                 "layers": ["CIMPATH=map/highways.json"]},
               "layerDefinitions": [layer_def]}
    project = {"type": "CIMGISProject", "version": "3.2.0", "projectItems": []}
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("GISProject.json", json.dumps(project))
        zf.writestr("Maps/a1b2c3.json", json.dumps(map_doc))
        zf.writestr("Index/thumbnail.png", b"\x89PNG\r\n\x1a\nnot really an image")
    Path(path).write_bytes(buf.getvalue())
    return path


def run(emit=None):
    emit = emit or print
    tmp = Path(tempfile.mkdtemp(prefix="tsmis_selftest_"))
    try:
        return _exercise(tmp, emit)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def _exercise(tmp, emit):
    emit("=" * 60)
    emit("TSMIS Branch Identifier -- bundle self-test")
    emit("=" * 60)
    import importlib
    for m in APP_MODULES:
        importlib.import_module(m)
    emit(f"modules import: {', '.join(APP_MODULES)} ok")
    import openpyxl
    emit(f"frozen={getattr(sys, 'frozen', False)}  openpyxl={openpyxl.__version__}")

    # 1. The real scan path over a synthetic project.
    from aprx_scan import run_scan
    from events import Events
    from scan_output import save
    projects = tmp / "projects"
    projects.mkdir()
    synthetic_aprx(projects / "Alpha.aprx", version="OWNER.Branch_A")
    (projects / "notes.txt").write_text("ignored", encoding="utf-8")
    lines = []
    result = run_scan(projects, recursive=True, events=Events(on_log=lines.append))
    assert len(result.projects) == 1, f"expected 1 project, got {len(result.projects)}"
    p = result.projects[0]
    assert p.status == "ok" and p.versions() == ["OWNER.Branch_A"], (p.status, p.versions(), p.message)
    workbook, diagnostics = save(result, tmp / "out")
    wb = openpyxl.load_workbook(str(workbook))
    ws = wb["Projects"]
    assert ws["D2"].value == "OWNER.Branch_A", ws["D2"].value
    diag = json.loads(diagnostics.read_text(encoding="utf-8"))
    assert diag["files"][0]["connections"][0]["version"] == "OWNER.Branch_A"
    emit(f"scan: 1 project -> {p.versions()} ; workbook + diagnostics written ok")

    # 2. GUI bridge: js_api + initial state + bundled ui/ assets.
    import webview
    import gui_api

    class _NoWorker:                  # no GitHub update check inside the gate
        def __init__(self, *a, **k):
            pass

        def start(self):
            pass

    gui_api.UpdateWorker = _NoWorker
    api = gui_api.GuiApi()
    state = api.get_initial_state()
    assert state["app_name"] and state["settings"]["values"], "GUI initial state incomplete"
    ui_index = gui_api._ui_index_path()
    assert ui_index.exists(), f"UI assets missing: {ui_index}"
    for asset in ("app.css", "app.js", "mock.js"):
        assert (ui_index.parent / asset).exists(), f"UI asset missing: {asset}"
    emit(f"gui: bridge api ok (ui={ui_index})")

    # 3. Hidden WebView window cycle — the ONLY skippable sub-check.
    res = {}

    def _drive(w):
        try:
            deadline = time.time() + 30
            while time.time() < deadline:
                if w.evaluate_js("typeof window.__tsmis !== 'undefined'"):
                    break
                time.sleep(0.25)
            res["state"] = w.evaluate_js("window.__tsmis.test_state()")
        except Exception as e:            # noqa: BLE001
            res["err"] = f"{type(e).__name__}: {e}"
        finally:
            w.destroy()

    window = webview.create_window("smoke", str(ui_index), js_api=api, hidden=True)
    api._window = window                  # the sender needs a window; attach() is for the real one
    window.events.loaded += lambda: threading.Thread(target=_drive, args=(window,),
                                                     daemon=True).start()
    watchdog = threading.Timer(60, lambda: (res.setdefault("err", "watchdog timeout"),
                                            window.destroy()))
    watchdog.daemon = True
    watchdog.start()
    try:
        webview.start(gui="edgechromium")
    except Exception as e:                # noqa: BLE001
        emit(f"gui: window skipped, environment can't start WebView2 ({type(e).__name__}: {e})")
    else:
        watchdog.cancel()
        if not res.get("state"):
            raise AssertionError(f"gui window cycle failed: {res.get('err', 'no JS state')}")
        emit(f"gui: WebView window + JS bridge ok ({res['state']})")
    emit("")
    emit("SELF-TEST OK -- every app-required code path works.")
    return 0
